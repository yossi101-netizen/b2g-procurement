"""
KritiKaal India→USA Landed-Cost Calculator — DEFINITIVE VERSION (v3)
Corrects all four flaws identified in the financial audit document:
1. EMS rates: discrete VLOOKUP lookup table (not linear CEILING)
2. GST on postal: applied via LUT toggle (Yes=0%, No=18%)
3. Postal duty: Section 122 + MFN (killed the 50% DDP ghost)
4. Weight cliff warning: alerts when near slab boundary
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, Protection,
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Calculator"

# ═══════════════════════════════════════════════════════════════
# ADMIN CONSTANT CELL MAP (single source of truth for references)
# Row 27 = zone header, Row 28 = sub-header, data starts row 29
# ═══════════════════════════════════════════════════════════════
FX       = "B29"   # INR per USD
SEC122   = "B30"   # Section 122 rate
MFN4205  = "B31"   # MFN HTS 4205
MFN6809  = "B32"   # MFN HTS 6809
BROK     = "B33"   # Brokerage USD
C_RATE   = "B34"   # Courier INR/kg
C_MIN    = "B35"   # Courier minimum INR
C_GST    = "B36"   # Courier GST
C_FUEL   = "B37"   # Courier fuel surcharge
VOLDIV   = "B38"   # Volumetric divisor
PACKBUF  = "B39"   # Default packing buffer %
PGST     = "B40"   # Postal freight GST

# EMS rate table rows 44-50 (row 42=header, 43=column labels)
EMS_RANGE = "$A$44:$B$50"
EMS_CEIL  = "$A$44:$C$50"
EMS_NEXT  = "$A$44:$D$50"

# ═══════════════════════════════════════════════════════════════
# STYLE DEFINITIONS
# ═══════════════════════════════════════════════════════════════
F_TITLE    = Font(name="Arial", size=14, bold=True, color="1A5276")
F_ZONE     = Font(name="Arial", size=11, bold=True, color="FFFFFF")
F_LABEL    = Font(name="Arial", size=10, color="2C3E50")
F_LABEL_B  = Font(name="Arial", size=10, bold=True, color="2C3E50")
F_INPUT    = Font(name="Arial", size=11, color="000000")
F_OUTPUT   = Font(name="Arial", size=10, bold=True, color="1A5276")
F_NOTE     = Font(name="Arial", size=9, italic=True, color="7F8C8D")
F_ADMIN    = Font(name="Arial", size=9, color="2C3E50")
F_ADMIN_H  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
F_WARN     = Font(name="Arial", size=10, bold=True, color="C0392B")
F_RESULT_W = Font(name="Arial", size=12, bold=True, color="FFFFFF")
F_SAVINGS  = Font(name="Arial", size=10, bold=True, color="1A5276")
F_POSTAL   = Font(name="Arial", size=11, bold=True, color="2E86C1")
F_COURIER  = Font(name="Arial", size=11, bold=True, color="E67E22")
F_SUB_HDR  = Font(name="Arial", size=8, bold=True, color="5D6D7E")
F_TABLE_H  = Font(name="Arial", size=9, bold=True, color="FFFFFF")
F_TABLE    = Font(name="Arial", size=9, color="2C3E50")
F_VERIFIED = Font(name="Arial", size=9, color="27AE60")
F_CLIFF    = Font(name="Arial", size=10, bold=True, color="D4AC0D")

FILL_WHITE  = PatternFill("solid", fgColor="FFFFFF")
FILL_OUTPUT = PatternFill("solid", fgColor="D6EAF8")
FILL_ADMIN  = PatternFill("solid", fgColor="F2F3F4")
FILL_TITLE  = PatternFill("solid", fgColor="EBF5FB")
FILL_ZONE_I = PatternFill("solid", fgColor="27AE60")
FILL_ZONE_O = PatternFill("solid", fgColor="2E86C1")
FILL_ZONE_A = PatternFill("solid", fgColor="5D6D7E")
FILL_ZONE_T = PatternFill("solid", fgColor="1A5276")
FILL_POSTAL = PatternFill("solid", fgColor="2E86C1")
FILL_COURIER= PatternFill("solid", fgColor="E67E22")
FILL_WARN   = PatternFill("solid", fgColor="FADBD8")
FILL_CLIFF  = PatternFill("solid", fgColor="FEF9E7")
FILL_TABLE  = PatternFill("solid", fgColor="EBF5FB")
FILL_TABLE2 = PatternFill("solid", fgColor="D6EAF8")

BDR = Border(
    left=Side("thin", "BDC3C7"), right=Side("thin", "BDC3C7"),
    top=Side("thin", "BDC3C7"), bottom=Side("thin", "BDC3C7"),
)
BDR_INPUT = Border(
    left=Side("medium", "2E86C1"), right=Side("medium", "2E86C1"),
    top=Side("medium", "2E86C1"), bottom=Side("medium", "2E86C1"),
)

AL = Alignment(horizontal="left", vertical="center")
AC = Alignment(horizontal="center", vertical="center")
AW = Alignment(horizontal="left", vertical="center", wrap_text=True)

LOCKED   = Protection(locked=True)
UNLOCKED = Protection(locked=False)

# Column widths
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 36

# ═══════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════
def sc(row, col, val, font=F_LABEL, fill=FILL_OUTPUT, prot=LOCKED,
       border=BDR, align=AL, nf=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font; c.fill = fill; c.protection = prot
    c.border = border; c.alignment = align
    if nf: c.number_format = nf
    return c

def zone_header(row, text, fill, height=24):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    sc(row, 1, text, F_ZONE, fill, LOCKED, BDR, AC)
    for c in range(2, 6):
        ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=c).protection = LOCKED
    ws.row_dimensions[row].height = height

def spacer(row):
    ws.row_dimensions[row].height = 6
    for c in range(1, 6):
        ws.cell(row=row, column=c).fill = FILL_WHITE
        ws.cell(row=row, column=c).protection = LOCKED

def fill_row(row, fill):
    for c in range(1, 6):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill; cell.protection = LOCKED; cell.border = BDR

# ═══════════════════════════════════════════════════════════════
# ROW 1 — TITLE
# ═══════════════════════════════════════════════════════════════
ws.merge_cells("A1:E1")
sc(1, 1, "KRITIKAAL — India to USA Landed-Cost Calculator  v3.0 (2026-06-22)",
   F_TITLE, FILL_TITLE, LOCKED, BDR, AC)
for c in range(2, 6):
    ws.cell(row=1, column=c).fill = FILL_TITLE
    ws.cell(row=1, column=c).protection = LOCKED
ws.row_dimensions[1].height = 32

# ═══════════════════════════════════════════════════════════════
# ROWS 2-10 — INPUT ZONE
# ═══════════════════════════════════════════════════════════════
zone_header(2, "INPUT ZONE — Enter product data below", FILL_ZONE_I)

inputs = [
    (3,  "Actual Weight",          0.8,    "kg",                       None),
    (4,  "Item Length (raw)",      30,     "cm",                       None),
    (5,  "Item Width (raw)",       20,     "cm",                       None),
    (6,  "Item Height (raw)",      25,     "cm",                       None),
    (7,  "FOB Value",              3.00,   "USD",                      '"$"#,##0.00'),
    (8,  "HTS Branch",             "4205", "(select from dropdown)",   None),
    (9,  "Packing Buffer Override","",     "% (blank = default 15%)",  None),
    (10, "LUT Registered?",        "No",  "(Yes = GST exempt export)", None),
]

for row, label, default, unit, nf in inputs:
    fill_row(row, FILL_WHITE)
    sc(row, 1, label, F_LABEL_B, FILL_WHITE, LOCKED, BDR, AL)
    sc(row, 2, default, F_INPUT, FILL_WHITE, UNLOCKED, BDR_INPUT, AC, nf)
    sc(row, 3, unit, F_NOTE, FILL_WHITE, LOCKED, BDR, AL)

# Data validations
dv_hts = DataValidation(type="list", formula1='"4205,6809"', allow_blank=False)
dv_hts.error = "Select 4205 (leather) or 6809 (plaster)"
dv_hts.errorTitle = "Invalid HTS"
ws.add_data_validation(dv_hts)
dv_hts.add(ws["B8"])

dv_lut = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
dv_lut.error = "Select Yes (LUT registered) or No (counter shipment)"
dv_lut.errorTitle = "LUT Status"
dv_lut.prompt = "Yes = GST-exempt (LUT filed)\nNo = 18% GST at counter"
dv_lut.promptTitle = "LUT Status"
ws.add_data_validation(dv_lut)
dv_lut.add(ws["B10"])

for r in [3, 4, 5, 6, 7]:
    dv = DataValidation(type="decimal", operator="greaterThan", formula1="0")
    dv.error = "Must be a positive number"
    dv.errorTitle = "Invalid Input"
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=r, column=2))

# ═══════════════════════════════════════════════════════════════
# ROW 11 — SPACER
# ═══════════════════════════════════════════════════════════════
spacer(11)

# ═══════════════════════════════════════════════════════════════
# ROWS 12-25 — OUTPUT ZONE
# ═══════════════════════════════════════════════════════════════
zone_header(12, "OUTPUT ZONE — Auto-calculated (do not edit)", FILL_ZONE_O)

# ── Row 13: Packing Buffer Applied ──
fill_row(13, FILL_OUTPUT)
sc(13, 1, "Packing Buffer Applied", F_LABEL, FILL_OUTPUT)
sc(13, 2, f'=IF(B9="",{PACKBUF},B9/100)', F_OUTPUT, FILL_OUTPUT, LOCKED, BDR, AC, "0%")

# ── Row 14: Packed Dimensions ──
fill_row(14, FILL_OUTPUT)
sc(14, 1, "Packed Dimensions (L x W x H)", F_LABEL, FILL_OUTPUT)
sc(14, 2, "=ROUND(B4*(1+B13),1)", F_OUTPUT, FILL_OUTPUT, LOCKED, BDR, AC)
sc(14, 3, "=ROUND(B5*(1+B13),1)", F_OUTPUT, FILL_OUTPUT, LOCKED, BDR, AC)
sc(14, 4, "=ROUND(B6*(1+B13),1)", F_OUTPUT, FILL_OUTPUT, LOCKED, BDR, AC)
sc(14, 5, "cm (buffered for packing)", F_NOTE, FILL_OUTPUT)

# ── Row 15: Volumetric Weight ──
fill_row(15, FILL_OUTPUT)
sc(15, 1, "Volumetric Weight (packed)", F_LABEL, FILL_OUTPUT)
sc(15, 2, f"=ROUND((B14*C14*D14)/{VOLDIV},2)", F_OUTPUT, FILL_OUTPUT, LOCKED, BDR, AC)
sc(15, 3, "kg", F_NOTE, FILL_OUTPUT)

# ── Row 16: Chargeable Weights ──
fill_row(16, FILL_OUTPUT)
sc(16, 1, "Chargeable Weight", F_LABEL, FILL_OUTPUT)
sc(16, 2, "=B3", F_OUTPUT, FILL_OUTPUT, LOCKED, BDR, AC)
sc(16, 3, "kg (postal = actual)", F_NOTE, FILL_OUTPUT)
sc(16, 4, "=MAX(B3,B15)", F_OUTPUT, FILL_OUTPUT, LOCKED, BDR, AC)
sc(16, 5, "kg (courier = max of actual/vol)", F_NOTE, FILL_OUTPUT)

# ── Row 17: Volumetric Penalty Warning ──
fill_row(17, FILL_OUTPUT)
sc(17, 1, "Volumetric Penalty?", F_LABEL_B, FILL_OUTPUT)
ws.merge_cells("B17:E17")
sc(17, 2,
   '=IF(B15>B3,'
   '"YES — courier bills "&TEXT(B15,"0.00")&"kg instead of "&TEXT(B3,"0.00")&"kg",'
   '"NO — actual weight governs")',
   F_OUTPUT, FILL_OUTPUT, LOCKED, BDR, AL)

# ── Row 18: EMS Base Rate (VLOOKUP) ──
fill_row(18, FILL_OUTPUT)
sc(18, 1, "EMS Base Rate (INR)", F_LABEL, FILL_OUTPUT)
sc(18, 2, f"=VLOOKUP(B3,{EMS_RANGE},2,TRUE)", F_OUTPUT, FILL_OUTPUT, LOCKED, BDR, AC, "#,##0")
sc(18, 3, "INR (from rate table below)", F_NOTE, FILL_OUTPUT)
sc(18, 5,
   f'="Slab capacity: "&TEXT(VLOOKUP(B3,{EMS_CEIL},3,TRUE)-B3,"0.000")&"kg free before next slab"',
   F_NOTE, FILL_OUTPUT, LOCKED, BDR, AL)

# ── Row 19: MFN Duty Rate ──
fill_row(19, FILL_OUTPUT)
sc(19, 1, "MFN Duty Rate (HTS)", F_LABEL, FILL_OUTPUT)
sc(19, 2, f'=IF(B8="4205",{MFN4205},IF(B8="6809",{MFN6809},0))',
   F_OUTPUT, FILL_OUTPUT, LOCKED, BDR, AC, "0%")

# ── Row 20: Sub-headers ──
fill_row(20, FILL_OUTPUT)
sc(20, 1, "", F_LABEL, FILL_OUTPUT)
sc(20, 2, "Freight", F_SUB_HDR, FILL_OUTPUT, LOCKED, BDR, AC)
sc(20, 3, "Duty", F_SUB_HDR, FILL_OUTPUT, LOCKED, BDR, AC)
sc(20, 4, "Brokerage", F_SUB_HDR, FILL_OUTPUT, LOCKED, BDR, AC)
sc(20, 5, "Total Landed (USD)", F_SUB_HDR, FILL_OUTPUT, LOCKED, BDR, AC)

# ── Row 21: POSTAL lane ──
fill_row(21, FILL_OUTPUT)
sc(21, 1, "POSTAL (India Post EMS)", F_POSTAL, FILL_OUTPUT)
sc(21, 2, f'=ROUND(B18*(1+IF(B10="Yes",0,{PGST}))/{FX},2)',
   F_OUTPUT, FILL_OUTPUT, LOCKED, BDR, AC, '"$"#,##0.00')
sc(21, 3, f"=ROUND(({SEC122}+B19)*B7,2)",
   F_OUTPUT, FILL_OUTPUT, LOCKED, BDR, AC, '"$"#,##0.00')
sc(21, 4, 0, F_NOTE, FILL_OUTPUT, LOCKED, BDR, AC, '"$"#,##0.00')
sc(21, 5, "=B21+C21", F_OUTPUT, FILL_OUTPUT, LOCKED, BDR, AC, '"$"#,##0.00')

# ── Row 22: COURIER lane ──
fill_row(22, FILL_OUTPUT)
sc(22, 1, "COURIER (Shiprocket/DHL/FedEx)", F_COURIER, FILL_OUTPUT)
sc(22, 2, f"=ROUND(MAX(D16*{C_RATE},{C_MIN})*(1+{C_GST}+{C_FUEL})/{FX},2)",
   F_OUTPUT, FILL_OUTPUT, LOCKED, BDR, AC, '"$"#,##0.00')
sc(22, 3, f"=ROUND(({SEC122}+B19)*B7,2)",
   F_OUTPUT, FILL_OUTPUT, LOCKED, BDR, AC, '"$"#,##0.00')
sc(22, 4, f"={BROK}", F_OUTPUT, FILL_OUTPUT, LOCKED, BDR, AC, '"$"#,##0.00')
sc(22, 5, "=B22+C22+D22", F_OUTPUT, FILL_OUTPUT, LOCKED, BDR, AC, '"$"#,##0.00')

# ── Row 23: OPTIMAL LANE ──
fill_row(23, FILL_OUTPUT)
ws.row_dimensions[23].height = 30
sc(23, 1, "OPTIMAL LANE", Font(name="Arial", size=12, bold=True, color="1A5276"), FILL_OUTPUT)
sc(23, 2, '=IF(E21<=E22,"POSTAL","COURIER")',
   F_RESULT_W, FILL_POSTAL, LOCKED, BDR, AC)
sc(23, 3, "=ROUND(ABS(E21-E22),2)", F_SAVINGS, FILL_OUTPUT, LOCKED, BDR, AC,
   '"saves $"#,##0.00')
sc(23, 4, "per parcel", F_NOTE, FILL_OUTPUT)

# ── Row 24: Weight Cliff Warning ──
fill_row(24, FILL_OUTPUT)
sc(24, 1, "Weight Cliff Alert", F_LABEL_B, FILL_OUTPUT)
ws.merge_cells("B24:E24")
sc(24, 2,
   f'=IF(AND(B3>=VLOOKUP(B3,{EMS_CEIL},3,TRUE)*0.95,'
   f'VLOOKUP(B3,{EMS_NEXT},4,TRUE)>0),'
   f'"WARNING: "&TEXT((VLOOKUP(B3,{EMS_CEIL},3,TRUE)-B3)*1000,"0")&'
   f'"g to next slab = +"&TEXT(VLOOKUP(B3,{EMS_NEXT},4,TRUE)-B18,"#,##0")&'
   f'" INR freight jump ($"&TEXT((VLOOKUP(B3,{EMS_NEXT},4,TRUE)-B18)*(1+IF(B10="Yes",0,{PGST}))/{FX},"0.00")&")",'
   f'"")',
   F_CLIFF, FILL_OUTPUT, LOCKED, BDR, AW)

# ── Row 25: GST Status Note ──
fill_row(25, FILL_OUTPUT)
sc(25, 1, "GST Status", F_LABEL, FILL_OUTPUT)
ws.merge_cells("B25:E25")
sc(25, 2,
   f'=IF(B10="Yes","LUT active — postal freight is GST-exempt (zero-rated export)",'
   f'"No LUT — 18% GST ("&TEXT({PGST}*100,"0")&"%) applied to postal freight at counter")',
   F_NOTE, FILL_OUTPUT, LOCKED, BDR, AL)

# ═══════════════════════════════════════════════════════════════
# ROW 26 — SPACER
# ═══════════════════════════════════════════════════════════════
spacer(26)

# ═══════════════════════════════════════════════════════════════
# ROWS 27-40 — ADMIN CONSTANTS ZONE
# ═══════════════════════════════════════════════════════════════
zone_header(27, "ADMIN CONSTANTS — Yossi-only (do not edit without authorization)", FILL_ZONE_A)

fill_row(28, FILL_ADMIN)
sc(28, 1, "Constant", Font(name="Arial", size=9, bold=True, color="5D6D7E"), FILL_ADMIN)
sc(28, 2, "Value", Font(name="Arial", size=9, bold=True, color="5D6D7E"), FILL_ADMIN, LOCKED, BDR, AC)
sc(28, 3, "Notes", Font(name="Arial", size=9, bold=True, color="5D6D7E"), FILL_ADMIN)

constants = [
    # (row, label, value, note, number_format)
    # These rows MUST match the cell map at the top of this script
    (29, "FX (INR per USD)",         86,   "Update monthly",                          "#,##0.00"),
    (30, "Section 122 Rate",         0.10, "EXPIRES Jul 23 2026",                    "0%"),
    (31, "MFN — HTS 4205",          0,    "Leather articles — Free",                 "0%"),
    (32, "MFN — HTS 6809",          0,    "Plaster articles — Free",                 "0%"),
    (33, "Brokerage (USD)",          15,   "Courier disbursement fee",                '"$"#,##0.00'),
    (34, "Courier Rate (INR/kg)",    1100, "REPLACE with reseller quote",             "#,##0"),
    (35, "Courier Minimum (INR)",    1449, "Floor charge",                            "#,##0"),
    (36, "Courier GST",             0.18,  "18% standard",                            "0%"),
    (37, "Courier Fuel Surcharge",   0,    "Set >0 if not GST-inclusive",             "0%"),
    (38, "Volumetric Divisor",       5000, "Air standard (5000 cm3/kg)",              "#,##0"),
    (39, "Default Packing Buffer %", 0.15, "15% dimensional uplift for packing",     "0%"),
    (40, "Postal Freight GST",       0.18, "18% if no LUT; auto-zeroed when LUT=Yes","0%"),
]

for row, label, value, note, nf in constants:
    fill_row(row, FILL_ADMIN)
    sc(row, 1, label, F_ADMIN, FILL_ADMIN)
    sc(row, 2, value, F_ADMIN, FILL_ADMIN, LOCKED, BDR, AC, nf)
    sc(row, 3, note, F_NOTE, FILL_ADMIN)

# ═══════════════════════════════════════════════════════════════
# ROW 41 — SPACER
# ═══════════════════════════════════════════════════════════════
spacer(41)

# ═══════════════════════════════════════════════════════════════
# ROWS 42-50 — EMS DISCRETE RATE LOOKUP TABLE
# ═══════════════════════════════════════════════════════════════
zone_header(42, "EMS RATE TABLE — India Post Merchandise to USA (Zone C)", FILL_ZONE_T)

fill_row(43, FILL_ZONE_T)
for col, hdr in [(1,"Weight From (kg)"), (2,"Rate (INR)"), (3,"Slab Ceiling (kg)"),
                 (4,"Next Slab Rate"), (5,"Status")]:
    sc(43, col, hdr, F_TABLE_H, FILL_ZONE_T, LOCKED, BDR, AC)

# Source: India Post Schedule of Rates 2025-26 via ClickPost EMS guide
# All 7 slabs verified from the same source
rate_table = [
    # (row, weight_from, rate_inr, ceiling_kg, next_rate, status)
    (44, 0,      1290,  0.500,  1720,  "VERIFIED"),
    (45, 0.501,  1720,  1.000,  2580,  "VERIFIED"),
    (46, 1.001,  2580,  2.000,  5010,  "VERIFIED"),
    (47, 2.001,  5010,  5.000,  8800,  "VERIFIED"),
    (48, 5.001,  8800,  10.000, 15500, "VERIFIED"),
    (49, 10.001, 15500, 20.000, 25200, "VERIFIED"),
    (50, 20.001, 25200, 30.000, 0,     "VERIFIED (max)"),
]

for i, (row, wf, rate, ceil_kg, next_r, status) in enumerate(rate_table):
    fill = FILL_TABLE if i % 2 == 0 else FILL_TABLE2
    fill_row(row, fill)
    sc(row, 1, wf,     F_TABLE, fill, LOCKED, BDR, AC, "0.000")
    sc(row, 2, rate,   F_TABLE, fill, LOCKED, BDR, AC, "#,##0")
    sc(row, 3, ceil_kg,F_TABLE, fill, LOCKED, BDR, AC, "0.000")
    sc(row, 4, next_r if next_r > 0 else "",
       F_TABLE, fill, LOCKED, BDR, AC, "#,##0")
    sc(row, 5, status, F_VERIFIED, fill, LOCKED, BDR, AC)

# ═══════════════════════════════════════════════════════════════
# CONDITIONAL FORMATTING
# ═══════════════════════════════════════════════════════════════
ws.conditional_formatting.add("B17:E17", FormulaRule(
    formula=['LEFT(B17,3)="YES"'], font=F_WARN, fill=FILL_WARN,
))
ws.conditional_formatting.add("B23", CellIsRule(
    operator="equal", formula=['"POSTAL"'],
    font=F_RESULT_W, fill=FILL_POSTAL,
))
ws.conditional_formatting.add("B23", CellIsRule(
    operator="equal", formula=['"COURIER"'],
    font=F_RESULT_W, fill=FILL_COURIER,
))
ws.conditional_formatting.add("B24:E24", FormulaRule(
    formula=['LEN(B24)>0'], font=F_CLIFF, fill=FILL_CLIFF,
))

# ═══════════════════════════════════════════════════════════════
# SHEET PROTECTION
# ═══════════════════════════════════════════════════════════════
ws.protection.sheet = True
ws.protection.password = "kritikaal2026"
ws.protection.formatCells = False
ws.protection.formatColumns = False
ws.protection.formatRows = False
ws.protection.sort = False
ws.protection.autoFilter = False

# ═══════════════════════════════════════════════════════════════
# PRINT / SAVE
# ═══════════════════════════════════════════════════════════════
ws.page_setup.orientation = "portrait"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.print_area = "A1:E50"

output = r"C:\Users\mygre\Desktop\3 - KritiKaal_Landed_Cost_Calculator.xlsx"
wb.save(output)
print(f"SAVED: {output}")
