"""
KritiKaal Leads Hunter — O-Output Layer
File: T-tools/exporter.py
Governed by: C-core/01-business-context.md, B-brain/01-tech-stack.md §7

Responsibilities (SRP boundary):
  - Query leads.db for ALL leads (Class A, B, C, Unclassified).
  - Aggregate signals across every classification run per lead (GROUP_CONCAT).
  - Write a three-tab Excel workbook (.xlsx) to the project root:
        Tab 1 — Class A        (QUALIFIED_A)
        Tab 2 — Class B        (QUALIFIED_B_PENDING_VERIFY)
        Tab 3 — Class C Review (DISQUALIFIED_C + UNCLASSIFIED)
  - Apply column-level formatting (bold headers, colour-coded tab tints).

This module does NOT:
  - Run any HTTP requests or NLP.
  - Write to leads.db (read-only with respect to the DB).
  - Filter or re-classify leads — DB status is the authoritative source.

Design decisions:
  - ALL records are exported so the human team can audit AI rejections (Class C tab).
  - Signals column shows the UNION of all signals from every classification run,
    not just the latest run (GROUP_CONCAT + Python merge/dedup).
  - Confidence column shows the MAX confidence ever recorded for the lead.
  - openpyxl is used instead of csv; utf-8 is the native encoding (no BOM needed).

Usage:
    cd T-tools && python exporter.py          # standalone
    from exporter import export_leads, DEFAULT_EXPORT_PATH   # from live_run.py
"""

from dotenv import load_dotenv
load_dotenv()

import json
import re
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

_HERE             = Path(__file__).resolve().parent
_ROOT             = _HERE.parent
_DB_PATH          = _HERE / "leads.db"
_EXPORT_PATH      = _ROOT / "kritikaal_leads_export.xlsx"
_RUN_HISTORY_PATH = _ROOT / "M-memory" / "run_history.json"

DEFAULT_DB_PATH     = str(_DB_PATH)
DEFAULT_EXPORT_PATH = str(_EXPORT_PATH)

# ---------------------------------------------------------------------------
# XML / openpyxl illegal character sanitizer
# ---------------------------------------------------------------------------

# Characters that openpyxl (and the underlying OOXML spec) refuse to encode.
# Mirrors the pattern in openpyxl.cell.cell.ILLEGAL_CHARACTERS_RE exactly so
# we strip precisely what the library rejects — no more, no less.
# Defined here rather than imported from a private openpyxl module so this
# stays stable across openpyxl versions.
#
# Excluded from XML 1.0:  \x00–\x08 (C0 controls except HT/LF/CR)
#                         \x0b–\x0c (VT, FF)
#                         \x0e–\x1f (remaining C0 controls)
_XLSX_ILLEGAL_CHARS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')


def _sanitize_cell(value) -> str:
    """
    Convert any cell value to a clean string safe for openpyxl.

    Two transforms applied in order:
      1. Cast to str (handles None, floats, ints uniformly).
      2. Strip XML-illegal control characters.

    Why here and not at classification time:
      The DB stores raw NLP output so the human team can audit it.
      Sanitizing at export time keeps the source data untouched while
      guaranteeing the Excel file is always writeable.

    Typical offender: regex pattern strings echoed by the LLM as signals
    (e.g. 'כריכ[הות]+\\s+עור') — these can contain backslash sequences
    that round-trip through JSON as control characters.
    """
    if value is None:
        return ""
    return _XLSX_ILLEGAL_CHARS_RE.sub("", str(value))


# ---------------------------------------------------------------------------
# Tab definitions — (tab_name, statuses_to_include, header_fill_hex)
# ---------------------------------------------------------------------------

_TABS: list[tuple[str, tuple[str, ...], str]] = [
    (
        "Class A — Volume Buyers",
        ("QUALIFIED_A",),
        "C6EFCE",    # light green
    ),
    (
        "Class B — Pending Verify",
        ("QUALIFIED_B_PENDING_VERIFY",),
        "FFEB9C",    # light yellow
    ),
    (
        "Class C — AI Rejections",
        ("DISQUALIFIED_C", "UNCLASSIFIED", "PENDING_LEGAL"),
        "FFC7CE",    # light red
    ),
]

# ---------------------------------------------------------------------------
# Column headers
# ---------------------------------------------------------------------------

COLUMNS: list[str] = [
    "Entity Name",          # 1
    "Domain",               # 2  — clickable hyperlink to website
    "WhatsApp",             # 3  — clickable wa.me link
    "Company ID (ח.פ)",    # 4
    "NLP Classification",   # 5
    "Confidence Score",     # 6
    "Top Signals",          # 7
    "Model / Tier",         # 8
    "Date Discovered",      # 9
    "Last Updated",         # 10
    "Source Agents",        # 11 — which pipeline agents discovered this lead
    "New This Run?",        # 12 — flagged GREEN for leads added in the latest run
]

_COL_WIDTHS: list[int] = [35, 22, 18, 14, 30, 10, 55, 14, 18, 18, 20, 13]

# ---------------------------------------------------------------------------
# Hyperlink & delta-flag styling
# ---------------------------------------------------------------------------

# Map internal vector names to human-readable agent labels.
_VECTOR_DISPLAY: dict[str, str] = {
    "V1_SERPER":              "Core Leather",
    "V1_LATERAL":             "Lateral",
    "V1_SHOPIFY_FINGERPRINT": "Shopify",
    "V1_COMPETITOR":          "Competitor",
    "V1_PROCUREMENT":         "Gov Procurement",
    "V1_SOCIAL_PAGE":         "Social-Only",
    "V1_PURCHASE_INTENT":     "Purchase Intent",
    "V2_SERPAPI":             "Maps/Local",
    "V3_DIRECT":              "Direct",
    "ig_signal_v1":           "Instagram",
}

# Font applied to hyperlinked cells (Excel convention: blue + underline).
_HYPERLINK_FONT = Font(color="0563C1", underline="single", size=10)

# Font and fill for "NEW" cells in the delta column.
_NEW_LEAD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_NEW_LEAD_FONT = Font(bold=True, color="375623", size=10)   # dark green text

# ---------------------------------------------------------------------------
# SQL — aggregate ALL classifications per lead across every run
#
# signals:    GROUP_CONCAT collects every signals JSON array ever returned
#             for this lead. Python merges + deduplicates them so no signal
#             discovered in an early run is silently dropped on re-runs.
# confidence: MAX across all runs (best score ever seen).
# model:      correlated subquery → most recent run's model name.
# ---------------------------------------------------------------------------

_QUERY_ALL = """
    SELECT
        l.entity_name,
        l.domain,
        l.whatsapp,
        l.company_id,
        l.status,
        MAX(lc.confidence)                       AS confidence,
        GROUP_CONCAT(lc.signals, '|||')          AS all_signals,
        (
            SELECT lc2.model_version
            FROM   lead_classifications lc2
            WHERE  lc2.lead_id = l.id
            ORDER  BY lc2.classified_at DESC
            LIMIT  1
        )                                        AS model_version,
        l.first_seen_at,
        l.last_updated_at,
        GROUP_CONCAT(DISTINCT ls.vector)         AS source_vectors
    FROM leads l
    LEFT JOIN lead_classifications lc ON lc.lead_id = l.id
    LEFT JOIN lead_sources        ls ON ls.lead_id = l.id
    GROUP BY l.id
    ORDER BY
        COALESCE(MAX(lc.confidence), 0.0) DESC,
        l.first_seen_at ASC
"""

# ---------------------------------------------------------------------------
# Field formatters
# ---------------------------------------------------------------------------

def _fmt_whatsapp(raw: Optional[str]) -> str:
    """972XXXXXXXXX → +972-XX-XXX-XXXX display format."""
    if not raw:
        return ""
    digits = raw.replace("+", "").replace("-", "").replace(" ", "")
    if digits.startswith("972") and len(digits) == 12:
        return f"+{digits[:3]}-{digits[3:5]}-{digits[5:8]}-{digits[8:]}"
    return raw


def _fmt_confidence(raw) -> str:
    """Float 0.0–1.0 → percentage string ('95%'). 'N/A' if None."""
    if raw is None:
        return "N/A"
    try:
        return f"{float(raw):.0%}"
    except (TypeError, ValueError):
        return "N/A"


def _fmt_signals(raw: Optional[str]) -> str:
    """
    Merge signals from ALL classification runs (GROUP_CONCAT result).

    raw format: '["sig1","sig2"]|||["sig2","sig3"]|||["sig1"]'
    Returns: unique signals joined by ' | ', top 5 only.
    """
    if not raw:
        return ""
    seen: dict[str, None] = {}    # insertion-order dict as ordered set
    for fragment in raw.split("|||"):
        fragment = fragment.strip()
        if not fragment:
            continue
        try:
            signals = json.loads(fragment)
            if isinstance(signals, list):
                for s in signals:
                    text = str(s).strip()
                    if text:
                        seen[text] = None
        except (json.JSONDecodeError, TypeError):
            continue
    return " | ".join(list(seen.keys())[:5])


def _fmt_vectors(raw: Optional[str]) -> str:
    """
    Format the GROUP_CONCAT(DISTINCT vector) result into human-readable labels.

    raw format: 'V1_SERPER,V1_LATERAL,V1_SHOPIFY_FINGERPRINT'  (comma-separated)
    Returns:    'Core Leather, Lateral, Shopify'

    Unknown vector values are passed through as-is so new agents are never
    silently dropped from the output.
    """
    if not raw:
        return ""
    parts = [v.strip() for v in raw.split(",") if v.strip()]
    labels = [_VECTOR_DISPLAY.get(p, p) for p in parts]
    return ", ".join(dict.fromkeys(labels))   # dedup while preserving order


def _fmt_datetime(raw: Optional[str]) -> str:
    """SQLite CURRENT_TIMESTAMP → 'YYYY-MM-DD HH:MM'."""
    if not raw:
        return ""
    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d %H:%M")
    except (ValueError, AttributeError):
        return str(raw)[:16]


def _fmt_status(raw: Optional[str]) -> str:
    _LABELS = {
        "QUALIFIED_A":                "Class A — Volume Buyer",
        "QUALIFIED_B_PENDING_VERIFY": "Class B — Pending Verify",
        "DISQUALIFIED_C":             "Class C — Disqualified",
        "UNCLASSIFIED":               "Unclassified",
        "PENDING_LEGAL":              "Pending Legal (no ח.פ)",
    }
    return _LABELS.get(raw or "", raw or "")


# ---------------------------------------------------------------------------
# Run-delta helpers — identify leads added in the most recent pipeline run
# ---------------------------------------------------------------------------

def _get_run_cutoff() -> Optional[str]:
    """
    Return the run_at ISO timestamp of the second-to-last pipeline run.

    Any lead with first_seen_at >= this value was added in the most recent
    run and should be flagged as "NEW" in the export.

    Returns None when:
      - The run_history.json file does not exist yet (fresh install).
      - Fewer than 2 runs have completed (all leads are implicitly "new").
      - The file cannot be parsed (graceful degradation).
    """
    if not _RUN_HISTORY_PATH.exists():
        return None
    try:
        history = json.loads(_RUN_HISTORY_PATH.read_text(encoding="utf-8"))
        if not isinstance(history, list) or len(history) < 2:
            return None   # first-ever run — all leads are new, no cutoff needed
        return history[-2]["run_at"]
    except (json.JSONDecodeError, KeyError, OSError):
        return None


def _get_new_domains(conn: sqlite3.Connection) -> frozenset:
    """
    Return the set of domains added in the most recent pipeline run.

    If no cutoff is available (first run or history file missing), returns
    ALL domains — every lead is treated as "new" on the first export.
    """
    cutoff = _get_run_cutoff()
    try:
        if cutoff is None:
            rows = conn.execute("SELECT domain FROM leads").fetchall()
        else:
            rows = conn.execute(
                "SELECT domain FROM leads WHERE first_seen_at >= ?", (cutoff,)
            ).fetchall()
        return frozenset(r[0] for r in rows if r[0])
    except Exception:
        return frozenset()


# ---------------------------------------------------------------------------
# Sheet builder
# ---------------------------------------------------------------------------

def _build_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    rows: list[sqlite3.Row],
    statuses: tuple[str, ...],
    fill_hex: str,
    new_domains: frozenset,
) -> int:
    """
    Write a single sheet: header row + data rows filtered by statuses.

    Special column treatment:
      Col 2  (Domain)        — clickable HYPERLINK to https://<domain>
      Col 3  (WhatsApp)      — clickable HYPERLINK to https://wa.me/<number>
      Col 11 (Source Agents) — human-readable agent labels from vector names
      Col 12 (New This Run?) — "NEW" with green fill for leads in new_domains

    Returns the number of data rows written.
    """
    header_fill = PatternFill(
        start_color=fill_hex, end_color=fill_hex, fill_type="solid"
    )
    header_font = Font(bold=True, size=10)
    cell_font   = Font(size=10)
    wrap_align  = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Header row
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align

    # Column widths + freeze header
    for col_idx, width in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    written = 0
    for row in rows:
        if row["status"] not in statuses:
            continue
        written += 1
        excel_row = written + 1   # +1 for the header

        domain      = row["domain"] or ""
        wa_raw      = row["whatsapp"]       # normalized 972XXXXXXXXX or None
        wa_display  = _fmt_whatsapp(wa_raw)
        is_new      = domain in new_domains

        data = [
            row["entity_name"]   or "",          # 1
            domain,                              # 2  — hyperlink applied below
            wa_display,                          # 3  — hyperlink applied below
            row["company_id"]    or "",          # 4
            _fmt_status(row["status"]),          # 5
            _fmt_confidence(row["confidence"]),  # 6
            _fmt_signals(row["all_signals"]),    # 7
            row["model_version"] or "N/A",       # 8
            _fmt_datetime(row["first_seen_at"]), # 9
            _fmt_datetime(row["last_updated_at"]),# 10
            _fmt_vectors(row["source_vectors"]), # 11
            "",                                  # 12 — overwritten by delta logic
        ]

        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(
                row=excel_row, column=col_idx,
                value=_sanitize_cell(value),
            )
            cell.font      = cell_font
            cell.alignment = wrap_align

            # ── Col 2: Domain — clickable link to website ──────────────────
            if col_idx == 2 and domain:
                cell.hyperlink = f"https://{domain}"
                cell.font      = _HYPERLINK_FONT

            # ── Col 3: WhatsApp — clickable wa.me link ─────────────────────
            elif col_idx == 3 and wa_raw:
                cell.hyperlink = f"https://wa.me/{wa_raw}"
                cell.font      = _HYPERLINK_FONT

            # ── Col 12: New This Run? — green badge for new leads ──────────
            elif col_idx == 12:
                if is_new:
                    cell.value     = "NEW"
                    cell.fill      = _NEW_LEAD_FILL
                    cell.font      = _NEW_LEAD_FONT
                    cell.alignment = center_align

    return written


# ---------------------------------------------------------------------------
# Public export function
# ---------------------------------------------------------------------------

def export_leads(
    db_path:     str = DEFAULT_DB_PATH,
    output_path: str = DEFAULT_EXPORT_PATH,
) -> dict[str, int]:
    """
    Query ALL leads from leads.db and write a three-tab Excel workbook.

    Tabs:
      1. Class A — Volume Buyers      (QUALIFIED_A)
      2. Class B — Pending Verify     (QUALIFIED_B_PENDING_VERIFY)
      3. Class C — AI Rejections      (DISQUALIFIED_C + UNCLASSIFIED + PENDING_LEGAL)

    All tabs are written regardless of row count (empty tabs are valid).
    Class C is included so the human team can audit AI disqualifications.

    Also removes the legacy .csv file if it exists (migration cleanup).

    Returns:
        dict with keys 'class_a', 'class_b', 'class_c', 'total' — row counts.
    """
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    all_rows   = conn.execute(_QUERY_ALL).fetchall()
    new_domains = _get_new_domains(conn)
    conn.close()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)    # remove the default empty sheet

    counts: dict[str, int] = {}
    count_keys = ("class_a", "class_b", "class_c")

    for (tab_name, statuses, fill_hex), key in zip(_TABS, count_keys):
        ws = wb.create_sheet(title=tab_name)
        n  = _build_sheet(ws, all_rows, statuses, fill_hex, new_domains)
        counts[key] = n

    counts["total"] = sum(counts.values())

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    # Migration cleanup — delete legacy .csv if it exists
    legacy_csv = Path(output_path).with_suffix(".csv")
    if legacy_csv.exists():
        legacy_csv.unlink()

    return counts


# ---------------------------------------------------------------------------
# Standalone entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import io
    import sys
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

    counts = export_leads()
    print("\n  O-Output Layer — Export complete (v2.0)")
    print(f"  Class A rows   : {counts['class_a']}")
    print(f"  Class B rows   : {counts['class_b']}")
    print(f"  Class C rows   : {counts['class_c']}")
    print(f"  Total rows     : {counts['total']}")
    print(f"  File           : {DEFAULT_EXPORT_PATH}\n")
